"""File upload service with security, storage, and content extraction.

Provides sandboxed file storage per project/agent with:
- File type validation (extension + magic bytes)
- Size limits and filename sanitization
- Path traversal prevention
- Content extraction for agent context
"""

import hashlib
import logging
import os
import re
import uuid
from dataclasses import dataclass
from pathlib import Path

from clara.config import settings

logger = logging.getLogger(__name__)

# Magic bytes for file type detection (first N bytes)
# More reliable than extension-only validation
FILE_SIGNATURES = {
    # PDF
    b"%PDF": "application/pdf",
    # Office Open XML (docx, xlsx, pptx) - ZIP-based
    b"PK\x03\x04": "application/zip",  # Will need further inspection
    # Old Office formats
    b"\xd0\xcf\x11\xe0": "application/msword",  # DOC, XLS, PPT (OLE2)
    # Images
    b"\x89PNG\r\n\x1a\n": "image/png",
    b"\xff\xd8\xff": "image/jpeg",
    b"GIF87a": "image/gif",
    b"GIF89a": "image/gif",
    b"RIFF": "image/webp",  # Will verify WEBP specifically
}

# Additional checks for ZIP-based formats
OOXML_CONTENT_TYPES = {
    b"word/": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    b"xl/": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    b"ppt/": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
}


@dataclass
class FileValidationResult:
    """Result of file validation."""
    is_valid: bool
    mime_type: str | None = None
    error_message: str | None = None
    checksum: str | None = None


@dataclass
class FileUploadResult:
    """Result of a file upload operation."""
    success: bool
    file_id: str | None = None
    storage_path: str | None = None
    stored_filename: str | None = None
    mime_type: str | None = None
    file_size: int = 0
    checksum: str | None = None
    extracted_text: str | None = None
    extraction_status: str | None = None
    error_message: str | None = None


class FileSecurityService:
    """Handles file validation and security checks."""

    # Dangerous filename patterns
    DANGEROUS_PATTERNS = [
        r"\.\.[\\/]",  # Path traversal
        r"^[\\/]",  # Absolute paths
        r"[<>:\"|?*]",  # Invalid characters
        r"[\x00-\x1f]",  # Control characters
        r"^(con|prn|aux|nul|com[0-9]|lpt[0-9])(\.|$)",  # Windows reserved names
    ]

    # Maximum filename length
    MAX_FILENAME_LENGTH = 200

    @classmethod
    def sanitize_filename(cls, filename: str) -> str:
        """Sanitize a filename to prevent security issues.

        Args:
            filename: The original filename

        Returns:
            Sanitized filename safe for storage
        """
        if not filename:
            return "unnamed_file"

        # Normalize unicode
        filename = filename.encode('utf-8', errors='ignore').decode('utf-8')

        # Get just the filename, not any path components
        filename = os.path.basename(filename)

        # Remove or replace dangerous characters
        for pattern in cls.DANGEROUS_PATTERNS:
            filename = re.sub(pattern, "_", filename, flags=re.IGNORECASE)

        # Replace spaces and other problematic chars
        filename = re.sub(r'[^\w.\-]', '_', filename)

        # Collapse multiple underscores
        filename = re.sub(r'_+', '_', filename)

        # Truncate if too long (preserve extension)
        if len(filename) > cls.MAX_FILENAME_LENGTH:
            name, ext = os.path.splitext(filename)
            max_name_len = cls.MAX_FILENAME_LENGTH - len(ext) - 1
            filename = name[:max_name_len] + ext

        # Ensure not empty
        if not filename or filename == ".":
            filename = "unnamed_file"

        return filename

    @classmethod
    def validate_file(
        cls,
        file_content: bytes,
        filename: str,
        max_size_bytes: int | None = None
    ) -> FileValidationResult:
        """Validate a file for security and compliance.

        Args:
            file_content: The raw file bytes
            filename: The original filename
            max_size_bytes: Maximum allowed size (uses config default if None)

        Returns:
            FileValidationResult with validation status and details
        """
        max_size = max_size_bytes or (settings.max_file_size_mb * 1024 * 1024)

        # Check file size
        if len(file_content) > max_size:
            return FileValidationResult(
                is_valid=False,
                error_message=f"File exceeds maximum size of {settings.max_file_size_mb}MB"
            )

        if len(file_content) == 0:
            return FileValidationResult(
                is_valid=False,
                error_message="File is empty"
            )

        # Check extension
        ext = os.path.splitext(filename)[1].lower()
        if ext not in settings.allowed_file_extensions:
            allowed = ', '.join(settings.allowed_file_extensions)
            return FileValidationResult(
                is_valid=False,
                error_message=f"File type '{ext}' is not allowed. Allowed types: {allowed}"
            )

        # Detect MIME type from file content (magic bytes)
        detected_mime = cls._detect_mime_type(file_content)
        if not detected_mime:
            # For text files, allow if extension is text-based
            if ext in ['.txt', '.md', '.csv']:
                # Verify it's actually text (no binary content)
                try:
                    file_content.decode('utf-8')
                    detected_mime = "text/plain" if ext == '.txt' else (
                        "text/markdown" if ext == '.md' else "text/csv"
                    )
                except UnicodeDecodeError:
                    return FileValidationResult(
                        is_valid=False,
                        error_message="File appears to be binary but has text extension"
                    )
            else:
                return FileValidationResult(
                    is_valid=False,
                    error_message="Could not verify file type from content"
                )

        # Verify detected MIME type is allowed
        if detected_mime not in settings.allowed_mime_types:
            return FileValidationResult(
                is_valid=False,
                error_message=f"Detected file type '{detected_mime}' is not allowed"
            )

        # Calculate checksum
        checksum = hashlib.sha256(file_content).hexdigest()

        return FileValidationResult(
            is_valid=True,
            mime_type=detected_mime,
            checksum=checksum
        )

    @classmethod
    def _detect_mime_type(cls, content: bytes) -> str | None:
        """Detect MIME type from file content using magic bytes.

        Args:
            content: File content bytes

        Returns:
            Detected MIME type or None if unknown
        """
        if len(content) < 8:
            return None

        # Check magic bytes
        for signature, mime_type in FILE_SIGNATURES.items():
            if content.startswith(signature):
                # Special handling for ZIP-based formats (OOXML)
                if mime_type == "application/zip":
                    return cls._detect_ooxml_type(content)
                # Special handling for WEBP (RIFF container)
                if signature == b"RIFF" and len(content) >= 12:
                    if content[8:12] == b"WEBP":
                        return "image/webp"
                    continue
                return mime_type

        return None

    @classmethod
    def _detect_ooxml_type(cls, content: bytes) -> str | None:
        """Detect specific OOXML format from ZIP content.

        Args:
            content: ZIP file content

        Returns:
            Specific OOXML MIME type or generic ZIP
        """
        # Look for directory markers in the ZIP content
        for marker, mime_type in OOXML_CONTENT_TYPES.items():
            if marker in content[:2000]:  # Check first 2KB
                return mime_type

        # Generic ZIP (not an Office document)
        return None


class FileStorageService:
    """Handles sandboxed file storage."""

    def __init__(self, base_path: str | None = None):
        """Initialize storage service.

        Args:
            base_path: Base directory for file storage (uses config default if None)
        """
        self.base_path = Path(base_path or settings.upload_dir)

    def get_project_path(self, project_id: str, agent_index: int) -> Path:
        """Get the sandboxed storage path for a project/agent.

        Args:
            project_id: The project ID
            agent_index: The agent index within the project

        Returns:
            Path to the project/agent storage directory
        """
        # Sanitize project_id to prevent path traversal
        safe_project_id = re.sub(r'[^\w\-]', '_', project_id)
        return self.base_path / safe_project_id / f"agent_{agent_index}"

    def ensure_directory(self, path: Path) -> None:
        """Ensure a directory exists with proper permissions.

        Args:
            path: Directory path to create
        """
        path.mkdir(parents=True, exist_ok=True)
        # Set restrictive permissions (owner only)
        os.chmod(path, 0o700)

    def store_file(
        self,
        file_content: bytes,
        project_id: str,
        agent_index: int,
        original_filename: str
    ) -> tuple[str, str]:
        """Store a file in the sandboxed location.

        Args:
            file_content: The file content to store
            project_id: Project ID for sandboxing
            agent_index: Agent index for sandboxing
            original_filename: Original filename (will be sanitized)

        Returns:
            Tuple of (stored_filename, relative_storage_path)
        """
        # Get sanitized filename with UUID prefix for uniqueness
        safe_filename = FileSecurityService.sanitize_filename(original_filename)
        unique_filename = f"{uuid.uuid4().hex[:8]}_{safe_filename}"

        # Get storage path
        storage_dir = self.get_project_path(project_id, agent_index)
        self.ensure_directory(storage_dir)

        # Write file
        file_path = storage_dir / unique_filename
        with open(file_path, 'wb') as f:
            f.write(file_content)

        # Set restrictive permissions
        os.chmod(file_path, 0o600)

        # Return relative path for database storage
        relative_path = str(file_path.relative_to(self.base_path))
        return unique_filename, relative_path

    def read_file(self, storage_path: str) -> bytes | None:
        """Read a file from storage.

        Args:
            storage_path: Relative path to the file

        Returns:
            File content or None if not found
        """
        file_path = self.base_path / storage_path

        # Security check: ensure path is within base_path
        try:
            file_path.resolve().relative_to(self.base_path.resolve())
        except ValueError:
            logger.warning(f"Path traversal attempt detected: {storage_path}")
            return None

        if not file_path.exists():
            return None

        with open(file_path, 'rb') as f:
            return f.read()

    def delete_file(self, storage_path: str) -> bool:
        """Delete a file from storage.

        Args:
            storage_path: Relative path to the file

        Returns:
            True if deleted, False otherwise
        """
        file_path = self.base_path / storage_path

        # Security check
        try:
            file_path.resolve().relative_to(self.base_path.resolve())
        except ValueError:
            logger.warning(f"Path traversal attempt in delete: {storage_path}")
            return False

        if file_path.exists():
            file_path.unlink()
            return True
        return False


class ContentExtractionService:
    """Extracts text content from files for agent context."""

    # Maximum extracted text length (50KB)
    MAX_EXTRACTED_LENGTH = 50000

    @classmethod
    def extract_text(cls, content: bytes, mime_type: str) -> tuple[str | None, str]:
        """Extract text content from a file.

        Args:
            content: File content bytes
            mime_type: The file's MIME type

        Returns:
            Tuple of (extracted_text, extraction_status)
            Status is one of: success, partial, failed, unsupported
        """
        try:
            if mime_type in ["text/plain", "text/markdown", "text/csv"]:
                return cls._extract_text_file(content)
            elif mime_type == "application/pdf":
                return cls._extract_pdf(content)
            elif mime_type in [
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "application/msword"
            ]:
                return cls._extract_docx(content)
            elif mime_type in [
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel"
            ]:
                return cls._extract_xlsx(content)
            elif mime_type.startswith("image/"):
                # Images don't have extractable text (would need OCR)
                return None, "unsupported"
            else:
                return None, "unsupported"
        except Exception as e:
            logger.exception(f"Error extracting content: {e}")
            return None, "failed"

    @classmethod
    def _extract_text_file(cls, content: bytes) -> tuple[str, str]:
        """Extract from plain text file."""
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                text = content.decode('latin-1')
            except Exception:
                return None, "failed"

        if len(text) > cls.MAX_EXTRACTED_LENGTH:
            return text[:cls.MAX_EXTRACTED_LENGTH], "partial"
        return text, "success"

    @classmethod
    def _extract_pdf(cls, content: bytes) -> tuple[str | None, str]:
        """Extract text from PDF file."""
        try:
            import pypdf
        except ImportError:
            logger.warning("pypdf not installed, cannot extract PDF content")
            return None, "unsupported"

        try:
            import io
            reader = pypdf.PdfReader(io.BytesIO(content))
            text_parts = []
            total_length = 0

            for page in reader.pages:
                page_text = page.extract_text() or ""
                if total_length + len(page_text) > cls.MAX_EXTRACTED_LENGTH:
                    # Truncate at limit
                    remaining = cls.MAX_EXTRACTED_LENGTH - total_length
                    text_parts.append(page_text[:remaining])
                    return "\n\n".join(text_parts), "partial"

                text_parts.append(page_text)
                total_length += len(page_text)

            full_text = "\n\n".join(text_parts)
            if full_text.strip():
                return full_text, "success"
            return None, "failed"

        except Exception as e:
            logger.warning(f"Failed to extract PDF: {e}")
            return None, "failed"

    @classmethod
    def _extract_docx(cls, content: bytes) -> tuple[str | None, str]:
        """Extract text from DOCX file."""
        try:
            from docx import Document
        except ImportError:
            logger.warning("python-docx not installed, cannot extract DOCX content")
            return None, "unsupported"

        try:
            import io
            doc = Document(io.BytesIO(content))
            text_parts = []
            total_length = 0

            for para in doc.paragraphs:
                para_text = para.text
                if total_length + len(para_text) > cls.MAX_EXTRACTED_LENGTH:
                    remaining = cls.MAX_EXTRACTED_LENGTH - total_length
                    text_parts.append(para_text[:remaining])
                    return "\n\n".join(text_parts), "partial"

                text_parts.append(para_text)
                total_length += len(para_text)

            full_text = "\n\n".join(text_parts)
            if full_text.strip():
                return full_text, "success"
            return None, "failed"

        except Exception as e:
            logger.warning(f"Failed to extract DOCX: {e}")
            return None, "failed"

    @classmethod
    def _extract_xlsx(cls, content: bytes) -> tuple[str | None, str]:
        """Extract text from XLSX file."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed, cannot extract XLSX content")
            return None, "unsupported"

        try:
            import io
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            text_parts = []
            total_length = 0

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"## Sheet: {sheet_name}")

                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if total_length + len(row_text) > cls.MAX_EXTRACTED_LENGTH:
                        return "\n".join(text_parts), "partial"

                    text_parts.append(row_text)
                    total_length += len(row_text) + 1

            full_text = "\n".join(text_parts)
            if full_text.strip():
                return full_text, "success"
            return None, "failed"

        except Exception as e:
            logger.warning(f"Failed to extract XLSX: {e}")
            return None, "failed"


class FileUploadService:
    """High-level service for file uploads with full pipeline."""

    def __init__(self):
        self.storage = FileStorageService()

    async def upload_file(
        self,
        file_content: bytes,
        filename: str,
        project_id: str,
        agent_index: int
    ) -> FileUploadResult:
        """Upload and process a file.

        Args:
            file_content: The file content
            filename: Original filename
            project_id: Project ID for sandboxing
            agent_index: Agent index for sandboxing

        Returns:
            FileUploadResult with upload status and details
        """
        # Step 1: Validate file
        validation = FileSecurityService.validate_file(file_content, filename)
        if not validation.is_valid:
            return FileUploadResult(
                success=False,
                error_message=validation.error_message
            )

        # Step 2: Store file
        try:
            stored_filename, storage_path = self.storage.store_file(
                file_content, project_id, agent_index, filename
            )
        except Exception as e:
            logger.exception("Failed to store file")
            return FileUploadResult(
                success=False,
                error_message=f"Failed to store file: {str(e)}"
            )

        # Step 3: Extract content
        extracted_text, extraction_status = ContentExtractionService.extract_text(
            file_content, validation.mime_type
        )

        # Step 4: Generate file ID
        file_id = f"file_{uuid.uuid4().hex[:16]}"

        return FileUploadResult(
            success=True,
            file_id=file_id,
            storage_path=storage_path,
            stored_filename=stored_filename,
            mime_type=validation.mime_type,
            file_size=len(file_content),
            checksum=validation.checksum,
            extracted_text=extracted_text,
            extraction_status=extraction_status
        )
